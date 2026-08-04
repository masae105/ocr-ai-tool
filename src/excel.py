import pandas as pd
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter


def save_to_excel(results, output_path):
    """
    請求書データをExcelへ保存する
    """

    invoice_rows = []
    detail_rows = []

    for data in results:

        # 請求書情報
        invoice_rows.append({
            "書類タイプ": data.get("書類タイプ"),
            "会社名": data.get("会社名"),
            "請求番号": data.get("請求番号"),
            "請求日": data.get("請求日"),
            "合計金額": int(data.get("合計金額", 0)),
            "金額チェック": data.get("金額チェック")
            
        })

        # 明細情報
        for item in data.get("明細", []):

            detail_rows.append({
                "商品名": item.get("商品名"),
                "金額": int(item.get("金額", 0))
            })


    invoice_df = pd.DataFrame(invoice_rows)
    detail_df = pd.DataFrame(detail_rows)


    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:


        invoice_df.to_excel(writer,sheet_name="請求書情報",index=False)


        detail_df.to_excel(writer,sheet_name="明細",index=False)


        invoice_ws = writer.sheets["請求書情報"]
        detail_ws = writer.sheets["明細"]


        # =========================
        # タイトル追加
        # =========================

        invoice_ws.insert_rows(1)

        invoice_ws.merge_cells("A1:F1")

        invoice_ws["A1"] = "請求書データ一覧"

        invoice_ws["A1"].font = Font(bold=True,size=14)

        invoice_ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        invoice_ws["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )


        # =========================
        # スタイル設定
        # =========================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="4F81BD"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center"
        )


        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )


        # =========================
        # ヘッダー装飾
        # =========================

        # タイトル追加後なので2行目
        for cell in invoice_ws[2]:

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


        for cell in detail_ws[1]:

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment



        # =========================
        # 罫線
        # =========================

        for ws in [invoice_ws, detail_ws]:

            for row in ws.iter_rows():

                for cell in row:

                    cell.border = thin_border



        # =========================
        # 列幅調整
        # =========================

        for ws in [invoice_ws, detail_ws]:

            for column in ws.columns:

                max_length = 0

                for cell in column:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )


                ws.column_dimensions[
                    get_column_letter(column[0].column)
                ].width = max_length + 4



        # 請求書情報シート幅固定

        invoice_ws.column_dimensions["A"].width = 18
        invoice_ws.column_dimensions["B"].width = 18
        invoice_ws.column_dimensions["C"].width = 15
        invoice_ws.column_dimensions["D"].width = 15
        invoice_ws.column_dimensions["E"].width = 15
        invoice_ws.column_dimensions["F"].width = 18



        # =========================
        # 表示形式
        # =========================

        # 合計金額
        for cell in invoice_ws["E"][2:]:

            cell.number_format = "#,##0"
            cell.alignment = Alignment(
                horizontal="right"
            )


        # 明細金額
        for cell in detail_ws["C"][1:]:

            cell.number_format = "#,##0"
            cell.alignment = Alignment(
                horizontal="right"
            )


        # 請求番号・日付中央寄せ

        for cell in invoice_ws["C"][2:]:

            cell.alignment = Alignment(
                horizontal="center"
            )


        for cell in invoice_ws["D"][2:]:

            cell.alignment = Alignment(
                horizontal="center"
            )


        for cell in detail_ws["A"][1:]:

            cell.alignment = Alignment(
                horizontal="center"
            )


        # =========================
        # 金額チェック色分け
        # =========================

        for row in range(
            3,
            invoice_ws.max_row + 1
        ):

            cell = invoice_ws[f"F{row}"]


            if cell.value == "OK":

                cell.font = Font(
                    color="008000",
                    bold=True
                )


            elif cell.value == "NG":

                cell.font = Font(
                    color="FF0000",
                    bold=True
                )